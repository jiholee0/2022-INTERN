from email.parser import Parser
from email import policy
from email.parser import BytesParser
import os
import email
from email.message import EmailMessage
from email.header import decode_header
import ctypes
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import configparser  # config.ini 파일을 불러와서 입출력 경로를 사용자가 설정할 수 있게 하기 위한 라이브러리
import datetime

# email내의 정보들을 추출하는 함수
def extract_info(target_eml):
    with open(target_eml, 'rb') as fp:
        # 데이터 담는 딕셔너리
        info_dic = dict()
        msg = BytesParser(policy=policy.default).parse(fp)
        EML_SUBJECT = str(msg['Subject']).split()[0]
        info_dic["Inquiry No"] = EML_SUBJECT
        EML_DATE = str(msg['Date'])

        try:
            for part in msg.walk():  # walk visits message
                type = part.get_content_type()
                if type == 'text/html':
                    EML_BODY = str(msg.get_body(preferencelist=('html')).get_content())
                    soup = BeautifulSoup(EML_BODY, "html.parser")
                    table = soup.find('table')
                    tbody = table.select_one('tbody')
                    trs = soup.select('tr')

                    firstTb = soup.find('table').select('table')[0]
                    firstTrs = firstTb.select('tr')

                    secondTb = soup.find('table').select('table')[1]
                    secondTrs = secondTb.select('tr')

                    thirdTb = soup.find('table').select('table')[2]
                    thirdTrs = thirdTb.select('tr')

                    # deliveryDate xxx
                    info_dic["Delivery Date"] = ""

                    # Inquiry Date
                    info_date = firstTrs[0].select('td')[3].text.replace("\xa0", "").replace('.', '-')
                    info_dic["Inquiry Date"] = info_date

                    # ref
                    # info_ref = thirdTrs[1].select('td')[1].text.split('(')[0].replace("\xa0", "")
                    info_ref = thirdTrs[1].select('td')[1].text.split('-').replace("\xa0", "")
                    info_dic["Ref"] = '-'.join([info_ref[0].replace("\xa0", ""),info_ref[1].replace("\xa0", ""),info_ref[2].replace("\xa0", "")])

                    #status
                    info_status = thirdTrs[2].select('td')[1].text.replace("\xa0", "")
                    info_dic["Status(DBIA)"] = info_status

                    # Units
                    # 읽은 Nation에 따라 main 함수에서 별도로 처리

                    # Client
                    info_client = firstTrs[2].select('td')[1].text.split('(')[0].replace("\xa0","")
                    final_client = info_client.split('*')[0]
                    info_dic["Client"] = final_client

                    # Subject
                    info_subject = thirdTrs[3].select('td')[1].text.replace("\xa0", "")
                    info_dic["Subject"] = info_subject

                    # Nation
                    info_nation = thirdTrs[4].select('td')[1].text.replace("\xa0", "")
                    info_dic["Nation"] = info_nation

                    # Address
                    info_address = thirdTrs[5].select('td')[1].text.replace("\xa0", "")
                    info_dic["ADDRESS"] = info_address

                    # pres
                    info_pres = thirdTrs[7].select('td')[1].text.replace("\xa0", "").replace("\xa0", "")
                    info_dic["PRES"] = info_pres

                    # Registration Number
                    info_registration_number = thirdTrs[9].select('td')[1].text.replace("\xa0", "")
                    info_dic["Registration Number"] = info_registration_number

                    # Vat Number
                    info_vat_number = thirdTrs[10].select('td')[1].text.replace("\xa0", "")
                    info_dic["Vat Number"] = info_vat_number

                    # Person incharge
                    info_person = thirdTrs[11].select('td')[2].text.replace("\xa0", "")
                    info_dic["Person in Charge"] = info_person

                    # Remark
                    info_remark = thirdTrs[14].select('td')[1].text.replace("\xa0", "")
                    info_dic["Remark"] = info_remark

                elif type == 'text/plain':
                    EML_BODY = str(msg.get_body(preferencelist=('plain')).get_content())

        except Exception as Error:
            print(Error)
            pass

    return info_dic


def main(email_dir, output_file_name, nation_code_path, output_dir):
    # 국가 코드표 불러오기
    # dataframeXlsx = pd.read_xlsx(nation_code_path, engine='openpyxl')
    nationList = []
    # 프로그램 실행 PC의 보안 문제로 인해 EXCEL 직접 접근 대신 TXT 파일의 데이터 불러오게끔 수정
    with open(nation_code_path, "rt", encoding='UTF8') as nationFile:
        while True:
            line = nationFile.readline()
            if not line:
                break
            nationStr = line.strip().split('\t')
            nationList.append(nationStr)

    count = 0
    for root, dirs, files in os.walk(email_dir):
        if output_file_name not in output_dir:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Delivery Date", "Inquiry No", "Inquiry Date", "Investigation Date"
                             , "Status(DBIA)", "Ref", "Units", "Client", "Subject", "Nation"
                             ,  "ADDRESS", "PRES", "TEL", "FAX", "E-mail", "Homepage"
                             , "Registration Number", "Vat Number", "Person in Charge", "Remark"
                          ])
            # 열 너비
            for cols in range(1,21):
                sheet.column_dimensions[get_column_letter(cols)].width = 13
            for rows in sheet["A1":"T1"]:
                for cell in rows:
                    # 셀 배경색
                    cell.fill = PatternFill(patternType="solid", fgColor="d3d3d3")
                    # 굵게
                    cell.font = Font(bold=True)
                    # 셀 테두리
                    side = Side("thin",color="000000")
                    cell.border = Border(left=side, right=side, top=side, bottom=side)
        wb.save(output_dir + output_file_name)
        excel = openpyxl.load_workbook(output_dir + output_file_name)

        for file in files:
            excel_ws = excel.active
            target = f"{root}\{file}"
            list_dic = extract_info(target)
            unit = ""
            nation_code = ""

            # 실제 코드
            # nation_code.xslx의 ref와 code가 일치하는 국가의 국가명 = nation, 유닛 = unit으로 한다.
            for nation in nationList:
                if int(list_dic["Ref"][0:3]) == nation[1]:
                    nation_code = nation[2]
                    if (len(nation) == 3):
                        unit = ''
                    else:
                        unit = nation[3]
                    break

            # *는 불러오지 않은 정보.
            # *"Delivery Date", "Inquiry No", "Inquiry Date", *"Investigation Date", *"Status(DBIA)", "Ref", "Units", "Client", "Subject", "Nation"
            # "Remark", "ADDRESS", "PRES", *"TEL", *"FAX", *"E-mail", *"Homepage", "Registration Number", "Vat Number", "Person in Charge"
            excel_ws.append(["", list_dic["Inquiry No"], list_dic["Inquiry Date"], "", "", list_dic["Ref"], unit,
                             list_dic["Client"], list_dic["Subject"], nation_code, list_dic["ADDRESS"], list_dic["PRES"], "", "", "", "",
                             list_dic["Registration Number"], list_dic["Vat Number"], list_dic["Person in Charge"], list_dic["Remark"]])
            excel.save(output_dir + output_file_name)
            count += 1

if __name__ == "__main__":
    properties = configparser.ConfigParser()    # 클래스 객체 생성
    properties.read('./config.ini')             # 파일 읽기
    PATH = properties["PATH"]                   # 섹션 선택

    # 결과 파일명을 현재 날짜로
    dateformat = "%Y%m%d"
    current = datetime.datetime.now()
    output_file_name = "오더변환파일_"+current.strftime(dateformat)+".xlsx"
    main(PATH["email_dir"], output_file_name, PATH["nation_code_path"], PATH["output_dir"])
